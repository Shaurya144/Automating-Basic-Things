# Chapter 12 — Excel Files

import openpyxl

# load workbook
wb = openpyxl.load_workbook("example.xlsx")

# select sheet
sheet = wb.active

# read cell
print(sheet["A1"].value)

# write cell
sheet["B1"] = "Hello"

# loop through rows
for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
    print(row)

# create new sheet
wb.create_sheet(title="NewSheet")

# save changes
wb.save("example.xlsx")
